from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QErrorMessage, QMessageBox
from interface import Ui_MainWindow
import sys
import pandas as pd
from openpyxl import load_workbook
from webster import compute_webster
import os
from create_sigs import replicate_sigs
from src.utils import *
import warnings
import logging
from tqdm import tqdm
import shutil

LOGGER = logging.getLogger(__name__)
LOGGER.setLevel(logging.DEBUG)
f = logging.Formatter("%(asctime)s-%(levelname)s: %(message)s")

class WebsterWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.ui.open_pushButton.clicked.connect(self.subarea_open)
        self.ui.start_pushButton.clicked.connect(self.start)
        self.ui.datos_pushButton.clicked.connect(self.create_datos)
        self.ui.sigs_pushButton.clicked.connect(self.multiply_sigs)

    def subarea_open(self):
        self.subarea_directory = QFileDialog.getExistingDirectory(self, "Open Directory")
        if self.subarea_directory:
            self.ui.subarea_lineEdit.setText(self.subarea_directory)
            logPath = os.path.join(self.subarea_directory, "logs")
            if not os.path.exists(logPath):
                os.mkdir(logPath)

            fh = logging.FileHandler(os.path.join(logPath, "Distro.log"), mode='w')
            fh.setFormatter(f)
            LOGGER.addHandler(fh)

    def start(self) -> None:
        #Get list of codes
        try:
            error_message = QErrorMessage(self)
            self.listCodes = get_codes(self.subarea_directory, error_message)
        except AttributeError as e:
            error_message = QErrorMessage(self)
            return error_message.showMessage("There is no subarea folder found")
        except Exception as e:
            error_message = QErrorMessage(self)
            return error_message.showMessage(str(e))

        try:
            maxValueRelations = float(self.ui.lineEdit.text())
        except Exception as e:
            print("Error: el valor máximo de la suma de relaciones de saturación no es un número decimal.")
            raise e

        #Get vehicle path:
        pathParts = self.subarea_directory.split("/")
        projectParts = pathParts[:-2]
        subareaName = pathParts[-1]
        projectParts = "\\".join(projectParts)

        subareaFolder = os.path.join(
            projectParts,
            "7. Informacion de Campo",
            subareaName,
        )

        excel_by_agent = {
            "Vehicular": {
                "Tipico": None,
                "Atipico": None,
            },
            "Peatonal": {
                "Tipico": None,
                "Atipico": None,
            }
        }
        
        #Lectura del excel de datos:
        path_datos = os.path.join(
            self.subarea_directory,
            "Program_Configuration.xlsx",
        )

        finalPath = os.path.join(
                self.subarea_directory,
                "Program_Results.xlsx",
                )

        try:
            duplicate_name_sheets(
                listCodes=self.listCodes,
                finalPath=finalPath,
            )
        except Exception as inst:
            raise inst

        try:
            wb_WEBSTER = load_workbook(finalPath, read_only=False, data_only=False)
        except Exception as inst:
            error_message = QErrorMessage(self)
            return error_message.showMessage("No se encontro el archivo de template WEBSTER.xlsx")

        #Starting code by code

        self.ui.progressBar.setMinimum(0)
        self.ui.progressBar.setMaximum(len(self.listCodes))

        for iteration, code in enumerate(self.listCodes):
            print(f"\n{f' Calculando intersección {code} ':#^{50}}")
            try:
                excel_by_agent, intervals = get_dict_by_agent(subareaFolder, excel_by_agent, code)
            except Exception as inst:
                raise inst
            
            try:
                dfPhases = pd.read_excel(path_datos, sheet_name=code, header=0, usecols="A:E", nrows=11).dropna()
                if dfPhases.empty:
                    print("Considerado no semaforizado: ", code)
                    continue

                dfTurns = pd.read_excel(path_datos, sheet_name=code, header=0, usecols="A:H", nrows=51, skiprows=27).dropna()
                dfTurns['Fase'] = dfTurns['Fase'].apply(process_elem)
                dfTurns = dfTurns[dfTurns['Considerar'] != 'NO']

                dfLanes = pd.read_excel(path_datos, sheet_name=code, header=0, usecols="J:N", nrows=51, skiprows=27).dropna()
                dfLanes["Destino.1"] = pd.to_numeric(dfLanes["Destino.1"], errors="coerce")
                dfLanes["Destino.1"] = dfLanes["Destino.1"].astype("Int64")
                dfLanes["Origen.1"] = dfLanes["Origen.1"].astype("Int64")
            except KeyError as inst:
                error_message = QErrorMessage(self)
                return error_message.showMessage("Posiblemente estes usando un template antiguo de 'Program_Configuration' revisar si tiene la columna CONSIDERAR.")
            except FileNotFoundError as inst:
                error_message = QErrorMessage(self)
                return error_message.showMessage("No se encontro el archivo 'Program_Configuration.xlsx'")
            except Exception as inst:
                error_message = QErrorMessage(self)
                return error_message.showMessage(str(inst))
            
            wbVehicleTipico = load_workbook(
                excel_by_agent["Vehicular"]["Tipico"],
                read_only=True,
                data_only=True,
            )

            wbVehicleAtipico = load_workbook(
                excel_by_agent["Vehicular"]["Atipico"],
                read_only=True,
                data_only=True,
            )
            
            wbPedestrianTipico = load_workbook(
                excel_by_agent["Peatonal"]["Tipico"],
                read_only=True,
                data_only=True,
            )

            for i in tqdm(range(13)):
                #Factores
                if i+1 == 1 or i+1 == 9:
                    factor = 0.30
                elif i+1 == 2 or i+1 == 8 or i+1 == 13:
                    factor = 0.50
                else:
                    factor = 1.00

                #Horas puntas para los intervalos
                if 1 <= i+1 <= 3:
                    interval = intervals[0]
                elif i+1 == 4:
                    interval = slice(8*4, 9*4)
                elif i+1 == 5:
                    interval = intervals[1]
                elif i+1 == 6:
                    interval = slice(14*4, 15*4)
                elif 7 <= i+1 <= 8:
                    interval = intervals[2]
                elif 9 <= i+1 <= 10:
                    interval = intervals[3]
                elif i+1 == 11:
                    interval = intervals[4]
                elif 12 <= i+1:
                    interval = intervals[5]

                #print(f"Cargando: {i+1}/13")
                try:
                    compute_webster(
                        wbVehicleTipico,
                        wbVehicleAtipico,
                        wbPedestrianTipico,
                        dfTurns, dfLanes, dfPhases, #Dataframes enviados
                        interval, factor, wb_WEBSTER[code], i, #<------ Aquí se le enviaría el sheet que le corresponde
                        LOGGER,
                        maxValueRelations,
                        )
                except Exception as inst:
                    print("Error: ", str(inst))
                    LOGGER.error(str(inst))
                    continue

            self.ui.progressBar.setValue(iteration+1)

        wb_WEBSTER.save(
            os.path.join(
                self.subarea_directory,
                "Program_Results.xlsx",
                )
            )
        wbVehicleTipico.close()
        wbVehicleAtipico.close() 
        wbPedestrianTipico.close()
        wb_WEBSTER.close()

        self.ui.label.setText("Done!")
        self.ui.progressBar.setValue(len(self.listCodes))
        print(f"{' CÁLCULOS FINALIZADOS ':#^{50}}")

    def create_datos(self) -> None: #Ready
        #Get list of codes
        try:
            error_message = QErrorMessage(self)
            self.listCodes = get_codes(self.subarea_directory, error_message)
        except AttributeError as e:
            error_message = QErrorMessage(self)
            return error_message.showMessage("There is no subarea folder found")
        except Exception as e:
            error_message = QErrorMessage(self)
            return error_message.showMessage(str(e))
        
        script_dir = os.path.dirname(os.path.abspath(__file__))
        origin_route = os.path.join(script_dir, 'tools', 'DATOS.xlsx')
        
        destiny_route = os.path.join(
            self.subarea_directory,
            "Program_Configuration.xlsx",
        )

        shutil.copy2(origin_route, destiny_route)

        try:
            duplicate2(
                listCodes=self.listCodes,
                finalPath=destiny_route,
            )
        except Exception as inst:
            error_message = QErrorMessage(self)
            return error_message.showMessage(str(inst))

        ##################
        # Inserting data #
        ##################

        try:
            data2excel(
                subareaFolder=self.subarea_directory,
                destiny_route=destiny_route,
                )
        except Exception as inst:
            error_message = QErrorMessage(self)
            return error_message.showMessage(str(inst))

        #Showing message
        info_message = QMessageBox(self)
        info_message.setIcon(QMessageBox.Information)
        info_message.setWindowTitle("Info")
        info_message.setText("Se ha creado el archivo 'Program_Configuration.xlsx' exitosamente")
        return info_message.show()

    def multiply_sigs(self) -> None:
        """ Create SIG from a unique sig file in ./Tipico/HPM folder. """
        try:
            replicate_sigs(self.subarea_directory) #NOTE: Designed to vissim version 24
        except Exception as inst:
            error_message = QErrorMessage(self)
            error_message.showMessage(str(inst))

        self.ui.label.setText("Copied and modified!")

def main():
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    app = QApplication(sys.argv)
    app.processEvents()
    window = WebsterWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
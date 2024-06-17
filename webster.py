from openpyxl import load_workbook
from unidecode import unidecode
import numpy as np
import pandas as pd
import math
import bisect
from pathlib import Path
import warnings
from unidecode import unidecode
from src.utils import *
import logging
from itertools import product

warnings.filterwarnings("ignore",category=FutureWarning)

###########################
# Default data and tables #
###########################

laneClassification = { #NOTE: Pueden afectar los cálculos, pero son conservadores. Estan por carril.
    "Local": 800,
    "Avenida": 1200,
    "Expresa": 1600,
    None: None,
}

min_green = {
    1: [6,11], #Nro. case: [min green, green used]
    2: [9,14],
    3: [14,15],
    4: [9,20],
    5: [12,25],
    6: [14,21],
    7: [17,25],
}

def _process_list(turns: list):
    newList = []

    for item in turns:
        if isinstance(item, str):
            newList.extend(map(int, item.split(',')))
        else:
            newList.append(item)
    return newList

def _apply_green(usedCase):
    return min_green[usedCase][1]

#UCP Values:
HEAVY_VEHICLES = [
    "OMNIBUS",
    "MICROBUS",
    "CAMIONETA RURAL",
    "BUS INTERPROVINCIAL",
    "CAMION",
    "TRAILER",
]

df_Evi = pd.DataFrame(index = [0,200,400,600,800,1000,1200], columns = [1,2,3])
DATA_IZQ = [[1.1,   1.1,    1.1 ],
            [2.5,   2.0,    1.8 ],
            [5.0,   3.0,    2.5 ],
            [10.0,  5.0,    4.0 ],
            [13.0,  8.0,    6.0 ],
            [15.0,  13.0,   10.0],
            [15.0,  15.0,   15.0],]

df_Evi.loc[[0,200,400,600,800,1000,1200], [1,2,3]] = DATA_IZQ

df_Evd = pd.DataFrame(index = [0,50,200,400,800],columns=["Equivalente"])
DATA_DER =[[1.18],
           [1.21],
           [1.32],
           [1.52],
           [2.14]]

df_Evd.loc[[0,50,200,400,800],["Equivalente"]] = DATA_DER

##########################
# Reading vehicular data #
##########################

def compute_webster(
        wbVehicleTipico: str,
        wbVehicleAtipico: str,
        wbPedestrianTipico: str,
        dfTurns: pd.DataFrame,
        dfLanes: pd.DataFrame,
        dfPhases: pd.DataFrame,
        interval: slice, #slice(28, 32)
        FACTOR: float,
        wb_WEBSTER: load_workbook,
        scenario: int,
        logger: logging,
        ) -> None:

    if scenario <= 7:
        wb = wbVehicleTipico
        logger.info("Escenario Típico")
    else:
        wb = wbVehicleAtipico
        logger.info("Escenario Atípico")

    #UCP VALUES
    ws = wb['Inicio']

    vehicle_types = [unidecode(row[0].value).upper() for row in ws[slice("AD4", "AD23")] if row[0].value != "n"]
    ucp_values = [row[0].value for row in ws[slice("AE4", "AE23")]][:len(vehicle_types)]

    heavy_vehicle_ucp = {}

    for vehicle, ucp in zip(vehicle_types, ucp_values):
        if vehicle in HEAVY_VEHICLES:
            heavy_vehicle_ucp[vehicle] = float(ucp)

    list_indices = []
    for i, vehicle in enumerate(vehicle_types):
        if vehicle in HEAVY_VEHICLES:
            list_indices.append(i)

    ####################################################
    # PERCENTAGE OF HEAVY VEHICLES FROM TOTAL VEHICLES #
    ####################################################

    try:
        array_flow, list_origin, list_destination, list_vr_name = flows(
            vehicle_types,
            wb,
            FACTOR,
            interval,
        )
    except Exception as e:
        raise e
    
    logger.debug("Flujos:\t\tPassed")

    array_flow_ucp = array_flow.copy()
    for veh_type in range(len(array_flow_ucp)):
        array_flow_ucp[veh_type] *= float(ucp_values[veh_type])

    origin_matrix = list(set(list_origin))
    #destiny_matrix = list(set(list_destination))

    #FIXME: No siempre los orígenes coinciden entre los conteos y los datos de Webster
    listTurns = _process_list(dfTurns["Origen"].unique().tolist())
    listTurns = list(set(listTurns))

    if origin_matrix != listTurns:
        print("Los accesos no coinciden entre los conteos y los datos de Webster")
        logger.error("Los accesos no coinciden entre los conteos y los datos de Webster")
        print(origin_matrix, dfTurns["Origen"].unique().tolist())
        logger.error(origin_matrix)
        logger.error(dfTurns["Origen"].unique().tolist())
        if scenario <= 7:
            print("FILE: ", wbVehicleTipico) #HACK: Solo por ahora
        else:
            print("FILE: ", wbVehicleAtipico) #HACK: Solo por ahora

    #################################################################
    # Reading pedestrian data and obtaining maximum pedestrian flow # 
    #################################################################

    max_ped_flow = pedestrian_flows(wbPedestrianTipico, interval)

    ###############
    # Compute Fhv #
    ###############

    access_percentage = {}
    for origin in origin_matrix: #NOTE: Puse que faltaba testearlo, creo que sí estaba.
        heavy_flows = [0 for _ in range(len(array_flow))]
        total_flows = 0
        for veh_type in range(len(array_flow)):
            for i in range(len(list_origin)):
                if origin == list_origin[i]:
                    total_flows += sum(array_flow[veh_type][:,i])
        
            if veh_type in list_indices:
                for j in range(len(list_origin)):
                    if origin == list_origin[j]:
                        heavy_flows[veh_type] += sum(array_flow[veh_type][:,j])
        percentage = [round(x/total_flows,4) if total_flows!= 0 else 0 for x in heavy_flows]
        access_percentage[origin] = percentage

    logger.info("Porcentaje de vehículos pesados:\tPassed")

    fhv_by_access = {}
    for access in access_percentage:
        denominator = [P*(ucp_value-1) for P, ucp_value in zip(access_percentage[access], ucp_values)]
        denominator = 100 + sum(denominator)
        fhv = 100 / denominator
        fhv_by_access[access] = round(fhv,4)

    logger.info("Flujos en UCP:\t\tPassed")

    ##########################
    # Computing Normal Flows #
    ##########################

    df_flows = pd.DataFrame(index = origin_matrix, columns = ['Directo', 'Izquierda', 'Derecha'])

    for origin in origin_matrix:
        for direction in ["Izquierda", "Derecha", "Directo"]:
            compute_flows(origin, dfTurns, direction, df_flows, array_flow)

    logger.info("Flujos normales:\tPassed")

    #######################
    # Computing ADE Flows #
    #######################

    #TODO: To consider U turns as Left turns.

    #print(dfTurns)

    df_flows_ade = pd.DataFrame(index = origin_matrix, columns = ['Origen', 'Directo', 'Izquierda', 'Derecha'])
    count = 0
    for access in fhv_by_access:
        #Directo (qd):
        if not dfTurns[(dfTurns["Origen"] == access) & (dfTurns["Giro"] == "Directo")].empty:
            qd = math.ceil(df_flows.at[access, "Directo"]/0.95/fhv_by_access[access]) #<--- FACTOR DE HORA PUNTA SE CONSIDERO 0.95
        else: qd = 0
        logger.info(f"{access} qd: " + str(qd))
        
        #Izquierda (qvi):
        listLeftIndexes = dfTurns[(dfTurns["Origen"] == access) & (dfTurns["Giro"] == "Izquierda")].index.tolist()
        if listLeftIndexes:
            checkProtected = True
            for leftIndex in listLeftIndexes:
                if dfTurns.at[leftIndex,'Protegido'] == False:
                    checkProtected = False
                    break
                
            if checkProtected:
                Evi = 1.05
            else: #Usa dfLanes
                num_opp_lines = dfLanes.at[dfTurns.at[access,"Destino.1"], "Carriles"]
                if num_opp_lines >= 3: num_opp_lines = 3
                oppositeList = dfTurns[dfTurns["Origen"] == access]["Origen Opuesto"].tolist()
                oppositeFlow = 0
                for oppositeIndex in oppositeList:
                    oppositeFlow += df_flows.at[oppositeIndex, "Directo"]
                serie = [0,200,400,600,800,1000,1200]
                index = bisect.bisect_left(serie, oppositeFlow)

                if index == 0:
                    Evi = 1.1
                elif index < 7:
                    before_value = df_Evi.at[serie[index-1], num_opp_lines]
                    after_value = df_Evi.at[serie[index], num_opp_lines]
                    #Three rules
                    Evi = after_value - (after_value - before_value)*(serie[index]-oppositeFlow)/(serie[index]-serie[index-1])
                else: #TODO: Check if everything is right
                    before_value    = df_Evi.at[serie[index-2],num_opp_lines]
                    after_value     = df_Evi.at[serie[index-1],num_opp_lines]
                    Evi = after_value - (after_value - before_value)*(serie[index-1]-oppositeFlow)/(serie[index-1]-serie[index-2])

                Evi = round(Evi, 2)
            qvi = math.ceil(df_flows.at[access, "Izquierda"]/0.95/fhv_by_access[access]*Evi)
        else:
            qvi = 0
        logger.info(f"{access} qvi: " + str(qvi))

        #Derecha (qvd):
        serie = [0,50,200,400,800]
        index = bisect.bisect_left(serie, max_ped_flow)
        if not dfTurns[(dfTurns["Origen"] == access) & (dfTurns["Giro"] == "Derecha")].empty:
            if index < 5:
                before_value = df_Evd.at[serie[index-1], "Equivalente"]
                after_value = df_Evd.at[serie[index], "Equivalente"]
                Evd = after_value - (after_value - before_value)*(serie[index]-max_ped_flow)/(serie[index]-serie[index-1])
            else:
                before_value = df_Evd.at[serie[index-2], "Equivalente"]
                after_value = df_Evd.at[serie[index-1], "Equivalente"]
                Evd = after_value - (after_value - before_value)*(serie[index-1]-max_ped_flow)/(serie[index-1]-serie[index-2])
            Evd = round(Evd, 2)
            qvd = math.ceil(df_flows.at[access, "Derecha"]/0.95/fhv_by_access[access]*Evd)
        else:
            qvd = 0
        logger.info(f"{access} qvd: " + str(qvd))

        #Asignación de flujos directos equivalentes
        df_flows_ade.at[count, "Origen"] = access
        df_flows_ade.at[count, "Directo"] = qd
        df_flows_ade.at[count, "Izquierda"] = qvi
        df_flows_ade.at[count, "Derecha"] = qvd
        count += 1

    ######################
    # LOST TIME BY CYCLE #
    ######################

    #Lost time by pedestrian and vehicles phases
    dfPhases["Ii"] = dfPhases["Ambar"] + dfPhases["Todo Rojo"]
    L = dfPhases["Ii"].sum()

    logger.info("Tiempo perdido s/peatones: " + str(L))

    #print("Solo A y RR = ", L)
    
    #Lost time by pedestrian green phase
    dfPhases["green"] = dfPhases.apply(lambda row: _apply_green(row["Caso"]) if row["Tipo"] == "P" else 0, axis=1)
    L += dfPhases["green"].sum()

    logger.info("Tiempo perdido c/peatones: " + str(L))

    #print("Incluye peatonal = ", L)

    #################################################
    # MAX RELATIONS CURRENT FLOW AND SATURATED FLOW #
    ################################################# 

    #Origin by phases
    phasesList = []
    for index, value in dfTurns.iterrows():
        for valor in value["Fase"]:
            phasesList.append(valor)
    
    phasesList = list(set(phasesList))

    phasesDict = {}
    for noPhase in phasesList:
        phasesByAccess = []
        for index, value in dfTurns.iterrows():
            for valuePhase in value["Fase"]:
                if valuePhase == noPhase:
                    phasesByAccess.append(value["Origen"])
            phasesByAccess = list(set(phasesByAccess))
        phasesDict[noPhase] = phasesByAccess

    #Maximum flows
    maxRelations_by_phase = {}
    for phase, originList in phasesDict.items():
        maximum_flows_ade = 0
        for _, row in df_flows_ade.iterrows():
            if row["Origen"] in originList:
                numerator = row["Directo"] + row["Izquierda"] + row["Derecha"]
                capacity = laneClassification[dfLanes[dfLanes["Origen.1"] == row['Origen']]["Clasificación"].unique().tolist()[0]] #NOTE: ¿Existe una forma más pythonic de hacerlo? Es un objeto tipo dataframe
                denominator = dfLanes[dfLanes["Origen.1"] == row['Origen']]["Carriles"].unique().tolist()[0]*capacity #NOTE: ¿Existe una forma más pythonic de hacerlo? Es un objeto tipo dataframe
                sum_flows = numerator / denominator
                if sum_flows > maximum_flows_ade:
                    maximum_flows_ade = sum_flows
        maxRelations_by_phase[phase] = round(maximum_flows_ade, 3)

    logger.info("Relaciones maxima de flujo por fase: " + str(maxRelations_by_phase))
    #print(maxRelations_by_phase)

    #####################
    # Cmin CRITIC ROUTE #
    #####################

    #Matriz con capacidades y suma

    dfLanes["Capacidad"] = dfLanes["Clasificación"].apply(lambda x: laneClassification[x])

    dfCritic = df_flows_ade.copy()
    dfCritic["Suma"] = dfCritic["Directo"] + dfCritic["Izquierda"] + dfCritic["Derecha"]
    dfCritic["Capacidad"] = np.nan

    for indexCritic, rowCritic in dfCritic.iterrows():
        for _, rowLanes in dfLanes.iterrows():
            if rowCritic["Origen"] == rowLanes["Origen.1"]:
                dfCritic.loc[indexCritic, "Capacidad"] = rowLanes["Capacidad"]

    dfCritic["Y"] = np.nan
    for indexCritic, rowCritic in dfCritic.iterrows():
        try:
            dfCritic.loc[indexCritic, "Y"] = rowCritic["Suma"] / rowCritic["Capacidad"]
        except:
            continue

    dfCritic["ti"] = dfCritic["Y"].apply(lambda x: 5+111*x) #NOTE: xpi = 0.9 and C = 100 seconds

    #Combinaciones de todas las ruats
    listPaths = []
    for listOrigins in phasesDict.values():
        listPaths.append(listOrigins)

    combinations = list(product(*listPaths)) #combinations: lista de todas las rutas posibles

    #Obtención del movimiento crítico con combinations y dfCritic
    maxValueCombo = 0
    finalList = None
    for combo in combinations:
        sumTime = 0
        indexListCombo = []
        for originValue in combo: #Sumar todo el valor del combo y comparar con otros
            for indexCritic, rowCritic in dfCritic.iterrows():
                if rowCritic["Origen"] == originValue:
                    sumTime += rowCritic["ti"]
                    indexListCombo.append(indexCritic)

        if sumTime > maxValueCombo:
            maxValueCombo = sumTime
            finalList = indexListCombo

    Y = 0
    for indexSelected in finalList:
        Y += dfCritic.loc[indexSelected, "Y"]

    ###########################
    # IMPRESION DE DATAFRAMES #
    ###########################

    # print("df_flows_ade: \n", df_flows_ade)
    # print("dfCritic: \n",dfCritic)
    print(maxRelations_by_phase)

    ###########################
    # COMPUTING OPTIMAL CYCLE #
    ###########################

    ws = wb_WEBSTER['WEBSTER']

    MAX_GREEN = dfPhases["green"].max()

    if sum(maxRelations_by_phase.values()) <= 0.80: 
        Cw = (1.5*L + 5) // (1 - sum(maxRelations_by_phase.values()))
        #Cmin = L/(1-sum(maxRelations_by_phase.values())) #NOTE: Ha sido cambiado por el método de la ruta crítica
        Cmin = 15/(1-Y)
        Cp = L/(1-1.1*sum(maxRelations_by_phase.values()))
        Ccruce = MAX_GREEN + L
        Cmax = 150 + MAX_GREEN
        ws.cell(17,2).value = "Los flujos no superan la capacidad, seleccionar el tiempo de ciclo de las opciones propuestas."
        logger.info("Los flujos no superan la capacidad")
    else:
        Cw = 0
        Cmin = 0
        Cp = 0
        Ccruce = MAX_GREEN + L
        Cmax = 150+MAX_GREEN
        ws.cell(17,2).value = "Los flujos superan la capacidad, se recomienda al simulador establecer el Tiempo de Ciclo."
        logger.info("Los flujos superan la capacidad de 0.80")

    #Tiempo de Ciclo
    ws.cell(scenario+2,3).value = Cmax
    ws.cell(scenario+2,4).value = Cw
    ws.cell(scenario+2,5).value = Cmin
    ws.cell(scenario+2,6).value = Cp
    ws.cell(scenario+2,7).value = Ccruce

    #Ingresar A y RR:
    for rowIndex, rowPhase in dfPhases.iterrows():
        ws.cell(scenario+2, 23+rowIndex*3).value = rowPhase["Ambar"]
        ws.cell(scenario+2, 24+rowIndex*3).value = rowPhase["Todo Rojo"]

    #Ingreso de Ys:
    for phaseNo, relation in maxRelations_by_phase.items():
        ws.cell(scenario+2, 10+phaseNo).value = relation #TODO: EDITAR NUEVAMENTE SI HAY.

    #Verdes peatonales:
    pedestrianGreenAccumulative = 0
    for rowIndex, rowPhase in dfPhases.iterrows():
        if rowPhase["Tipo"] == "P":
            pedestrianGreen = min_green[rowPhase["Caso"]][1]
            ws.cell(scenario+2, 22+rowIndex*3).value = pedestrianGreen 
            pedestrianGreenAccumulative += pedestrianGreen

    #Pérdidas por verdes peatonales:
    ws.cell(scenario+2, 9).value = pedestrianGreenAccumulative

    #Tiempos mínimos de verde usados:
    for rowIndex, rowPhase in dfPhases.iterrows():
        minGreenGeneral = min_green[rowPhase["Caso"]][1]
        ws.cell(scenario+2, 37+rowIndex).value = minGreenGeneral

    return None
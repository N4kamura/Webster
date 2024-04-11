""" 
Factor de equivalencia de VEHÍCULOS PESADOS:
fhv = 100 / (100 + Pt(Et-1) + Pb(Eb-1) + Pr(Er-1))
fhv = factor de ajuste por efecto de los vehículos pesados
Pt = % de camiones en la corriente vehicular
Pb = % de buses en la corriente vehicular
Pr = % de vehículos recreativos en la corriente vehicular
Et = automóviles equivalentes a un camión
Eb = automóviles equivalentes a un bus
Er = automóviles equivalentes a un vehículo recreativo
"""

""" 
Hora punta de la madrugada      05:00 - 06:30 
Hora valle de la madrugada      06:30 - 10:30 
Hora punta de la mañana         10:30 - 12:30 
Hora valle de la mañana         12:30 - 15:00 
Hora punta de la tarde          15:00 - 17:00 
Hora valle de la tarde          17:00 - 22:00 
Hora punta de la noche          22:00 - 00:00 
Hora Valle Madrugada (HVMAD)    00:00 - 06:00 
Hora Punta Mañana (HPM)         06:00 - 12:00 
Hora Punta Tarde (HPT)          12:00 - 17:00 
Hora Punta Noche (HPN)          17:00 - 22:00 
Hora Valle Noche (HVN)          22:00 - 00:00 
"""

from openpyxl import load_workbook
from unidecode import unidecode
import numpy as np
import pandas as pd
import math
import bisect
from pathlib import Path
import warnings

warnings.filterwarnings("ignore",category=DeprecationWarning)

###########################
# Default data and tables #
###########################

min_green = {
    1: [6,11], #Nro. case: [min green, green used]
    2: [9,14],
    3: [14,15],
    4: [9,20],
    5: [12,25],
    6: [14,21],
    7: [17,25],
}

rr_time = {
    1: 2, #Nro.case: total red time
    2: 2,
    3: 2,
    4: 3,
    5: 3,
    6: 4,
    7: 4,
}

#UCP Values:
HEAVY_VEHICLES = [
    "OMNIBUS",
    "MICROBUS",
    "CAMIONETA RURAL",
    "BUS INTERPROVINCIAL",
    "CAMION",
    "TRAILER",
]

df_Evi = pd.DataFrame(index = [0,200,400,600,800,1000,1200], columns = [1,2,3])
DATA_IZQ = [[1.1,   1.1,    1.1 ],
            [2.5,   2.0,    1.8 ],
            [5.0,   3.0,    2.5 ],
            [10.0,  5.0,    4.0 ],
            [13.0,  8.0,    6.0 ],
            [15.0,  13.0,   10.0],
            [15.0,  15.0,   15.0],]

df_Evi.loc[[0,200,400,600,800,1000,1200], [1,2,3]] = DATA_IZQ

df_Evd = pd.DataFrame(index = [0,50,200,400,800],columns=["Equivalente"])
DATA_DER =[[1.18],
           [1.21],
           [1.32],
           [1.52],
           [2.14]]

df_Evd.loc[[0,50,200,400,800],["Equivalente"]] = DATA_DER

##########################
# Reading vehicular data #
##########################

def compute_webster(
        path_vehicles: str | Path,
        path_pedestrian: str | Path,
        min_green_id: int,
        rr_time_id: int,
        interval: slice, #slice(28, 32)
        df: dict,
        FACTOR: float,
        wb_WEBSTER: load_workbook,
        ) -> None:

    wb = load_workbook(path_vehicles, read_only=True, data_only=True)

    #UCP VALUES
    ws = wb['Inicio']

    vehicle_types = [unidecode(row[0].value).upper() for row in ws[slice("AD4", "AD23")] if row[0].value != "n"]
    ucp_values = [row[0].value for row in ws[slice("AE4", "AE23")]][:len(vehicle_types)]

    heavy_vehicle_ucp = {}

    for vehicle, ucp in zip(vehicle_types, ucp_values):
        if vehicle in HEAVY_VEHICLES:
            heavy_vehicle_ucp[vehicle] = float(ucp)

    list_indices = []
    for i, vehicle in enumerate(vehicle_types):
        if vehicle in HEAVY_VEHICLES:
            list_indices.append(i)

    ####################################################
    # PERCENTAGE OF HEAVY VEHICLES FROM TOTAL VEHICLES #
    ####################################################

    origin_slices = [
        slice("E12", "E21"),
        slice("K12", "K21"),
        slice("E24", "E33"),
        slice("K24", "K33"),
    ]

    destiny_slices = [
        slice("F12", "F21"),
        slice("L12", "L21"),
        slice("F24", "F33"),
        slice("L24", "L33"),
    ]

    giro_slices = [
        slice("G12", "G21"),
        slice("M12", "M21"),
        slice("G24", "G33"),
        slice("M24", "M33"),
    ]

    num_giros = [0,0,0,0]
    #7-8
    #interval = slice(28, 32) #TODO: <-------------------------------------------------------------------- AUTOMATIZAR PARA CADA ESCENARIO :D LOS VOLUMENES
    num_veh_classes = len(vehicle_types)

    for j, turn in enumerate(giro_slices):
        aux = []
        for row in ws[turn]:
            aux.append(row[0].value)
        try:
            quant = aux.index(None)
        except ValueError:
            quant = len(aux)
        num_giros[j] = quant

    hojas = ["N", "S", "E", "O"]

    list_destination    = []
    list_origin         = []
    list_vr_name        = []

    list_flow = []
    for i_giro in range(len(hojas)):
        ws = wb["Inicio"]
        slice_origin = origin_slices[i_giro]
        slice_destiny = destiny_slices[i_giro]
        slice_giros = giro_slices[i_giro]
        num_giro_i = num_giros[i_giro]

        list_origin.extend(
            [row[0].value for row in ws[slice_origin]][:num_giro_i]
        )

        list_destination.extend(
            [row[0].value for row in ws[slice_destiny]][:num_giro_i]
        )

        list_vr_name.extend(
            [row[0].value for row in ws[slice_giros]][:num_giro_i]
        )

        ws = wb[hojas[i_giro]]

        list_A = [[cell.value for cell in row] for row in ws["K16":"HB111"]]
        A = np.array(list_A, dtype="float")
        A[np.isnan(A)] = 0
        A = A*FACTOR

        list_flow.append(
            np.array(
                [
                    A[interval, (10*veh_type):(10*veh_type+num_giro_i)]
                    for veh_type in range(num_veh_classes)
                ]
            )
        )

        array_flow = np.concatenate(list_flow, axis=-1)

    wb.close()

    array_flow_ucp = array_flow.copy()
    for veh_type in range(len(array_flow_ucp)):
        array_flow_ucp[veh_type] *= float(ucp_values[veh_type])

    origin_matrix = list(set(list_origin))

    if origin_matrix != df.index.tolist():
        print("Los accesos no coinciden entre los conteos y los datos de Webster")

    ###########################
    # Reading pedestrian data #
    ###########################

    wb = load_workbook(path_pedestrian, read_only=True, data_only=True)

    ws = wb['Inicio']

    movimiento_slices = [
        slice("G13", "G22"),
        slice("K13", "K22"),
        slice("G25", "G34"),
        slice("K25", "K34"),
    ]

    types_ped_slices = [
        slice("W4", "W8"),
        slice("Y4", "Y8"),
        slice("AA4", "AA8"),
    ]

    num_ped_classes = 0
    for type_ped_slice in types_ped_slices:
        num_ped_classes += [row[0].value for row in ws[type_ped_slice]].index(None)

    num_giros = [0,0,0,0]
    for j, turn in enumerate(movimiento_slices):
        aux = []
        for row in ws[turn]:
            aux.append(row[0].value)
        try: quant = aux.index(None)
        except ValueError: quant = len(aux)
        num_giros[j] = quant

    list_movimientos = []
    list_ped = []

    for i_giro in range(len(num_giros)):
        ws = wb['Inicio']
        slice_movimiento = movimiento_slices[i_giro]
        num_giro_i = num_giros[i_giro]

        list_movimientos.extend([row[0].value for row in ws[slice_movimiento]][:num_giro_i])

        ws = wb['Data Peatonal']

        list_A = [[cell.value for cell in row] for row in ws["L20":"UY83"]]
        A = np.array(list_A, dtype="float")
        A[np.isnan(A)] = 0
        A = np.concatenate((np.zeros((24, A.shape[1])), A), axis = 0)

        list_ped.append(
            np.array(
                [
                    A[interval, (10*ped_type + 140*i_giro):(10*ped_type + num_giro_i + 140*i_giro)]
                    for ped_type in range(num_ped_classes)
                ]
            )
        )

        array_ped_flow = np.concatenate(list_ped, axis=-1)

    wb.close()

    ###########################
    # Maximum pedestrian flow #
    ###########################

    str_movements = [str(no) for no in list_movimientos]

    list_sums = []
    for ped_type in range(num_ped_classes):
        check_status = [False for _ in range(len(str_movements))]
        ped_flow = array_ped_flow[ped_type]
        list_sum_peds = []
        for i, giro in enumerate(str_movements):
            flow_cross = 0
            if check_status[i]: continue
            check_status[i] = True
            flow_cross += sum(ped_flow[:,i])
            for j, giro_inv in enumerate(str_movements):
                if giro[::-1] == giro_inv:
                    check_status[j] = True
                    flow_cross += sum(ped_flow[:,j])
                    break
            list_sum_peds.append(flow_cross)
        list_sums.append(list_sum_peds)

    sum_by_peds = [sum(x) for x in zip(*list_sums)]

    max_ped_flow = max(sum_by_peds)

    ###############
    # Compute Fhv #
    ###############

    access_percentage = {}
    for origin in origin_matrix: #Test this
        heavy_flows = [0 for _ in range(len(array_flow))]
        total_flows = 0
        for veh_type in range(len(array_flow)):
            for i in range(len(list_origin)):
                if origin == list_origin[i]:
                    total_flows += sum(array_flow[veh_type][:,i])
        
            if veh_type in list_indices:
                for j in range(len(list_origin)):
                    if origin == list_origin[j]:
                        heavy_flows[veh_type] += sum(array_flow[veh_type][:,j])
        percentage = [round(x/total_flows,4) if total_flows!= 0 else 0 for x in heavy_flows]
        access_percentage[origin] = percentage

    fhv_by_access = {}
    for access in access_percentage:
        denominator = [P*(ucp_value-1) for P, ucp_value in zip(access_percentage[access], ucp_values)]
        denominator = 100 + sum(denominator)
        fhv = 100 / denominator
        fhv_by_access[access] = round(fhv,4)

    """ data = {
        'Directo':     [2,1,4,3],
        'Izquierda':         [3,4,2,1],
        'Protegido':    [False, False, False, False],
        'Derecha':        [4,3,1,2],
        'Carriles':        [2,2,2,2],
        'Grupo':        [1,1,2,2],
    } """

    df_flows = pd.DataFrame(index = origin_matrix, columns = ['Directo', 'Izquierda', 'Derecho'])

    #df = pd.DataFrame(data, index = origin_matrix)

    for origin in origin_matrix:
        for destiny in origin_matrix:
            flow = 0
            if df.at[origin, "Izquierda"] == destiny:
                for (i, o), d in zip(enumerate(list_origin), list_destination): #Encuentra que giro es, el orden, no el tipo vehicular
                    if o == origin and d == destiny:
                        for veh_type in range(len(array_flow)): #Accade a los conteos de cada giro solo por un tipo vehicular
                            for giro in range(len(list_origin)): #Aquí accede al giro correspondiente para un tipo vehicular específico
                                if giro == i:
                                    flow += sum(array_flow[veh_type][:,giro]) #Aquí esta sumando de cada tipo vehicular un giro en específico
                    df_flows.at[origin, 'Izquierda'] = flow

            flow = 0
            if df.at[origin, "Derecha"] == destiny:
                for (i, o), d in zip(enumerate(list_origin), list_destination):
                    if o == origin and d == destiny:
                        for veh_type in range(len(array_flow)):
                            for giro in range(len(list_origin)):
                                if giro == i:
                                    flow += sum(array_flow[veh_type][:,giro])
                    df_flows.at[origin, 'Derecha'] = flow


            flow = 0
            if df.at[origin, "Directo"] == destiny:
                for (i, o), d in zip(enumerate(list_origin), list_destination):
                    if o == origin and d == destiny:
                        for veh_type in range(len(array_flow)):
                            for giro in range(len(list_origin)):
                                if giro == i:
                                    flow += sum(array_flow[veh_type][:,giro])
                    df_flows.at[origin, 'Directo'] = flow

    #######################
    # Computing ADE Flows #
    #######################

    df_flows_ade = pd.DataFrame(index = origin_matrix, columns = ['Directo', 'Izquierda', 'Derecha'])
    for access in fhv_by_access:
        #Directo:
        qd = math.ceil(df_flows.at[access, "Directo"]/0.95/fhv_by_access[access]) #<--- FACTOR DE HORA PUNTA SE CONSIDERO 0.95
        
        #Izquierda:
        if df.at[access,'Protegido']:
            Evi = 1.05
        else:
            num_opp_lines = df.at[df.at[access,'Directo'],'Carriles']
            if num_opp_lines >= 3: num_opp_lines = 3
            opposite_flow = df_flows.at[df.at[access,'Directo'],'Directo']
            serie = [0,200,400,600,800,1000,1200]
            index = bisect.bisect_left(serie, opposite_flow)
            if index == 0: Evi = 1.1
            before_value    = df_Evi.at[serie[index-1],num_opp_lines]
            after_value     = df_Evi.at[serie[index],num_opp_lines]
            Evi = after_value - (after_value - before_value)*(serie[index]-opposite_flow)/(serie[index]-serie[index-1])
            Evi = round(Evi, 2)
            #print(f"Opposite flow = {opposite_flow}, Number of Carriles = {num_opp_lines}, Evi = {Evi}")

        qvi = math.ceil(df_flows.at[access, "Izquierda"]/0.95/fhv_by_access[access]*Evi)

        #Derecha:
        serie = [0,50,200,400,800]
        index = bisect.bisect_left(serie, max_ped_flow)
        before_value = df_Evd.at[serie[index-1], "Equivalente"]
        after_value = df_Evd.at[serie[index], "Equivalente"]
        Evd = after_value - (after_value - before_value)*(serie[index]-max_ped_flow)/(serie[index]-serie[index-1])
        Evd = round(Evd, 2)

        qvd = math.ceil(df_flows.at[access, "Derecha"]/0.95/fhv_by_access[access]*Evd)

        df_flows_ade.at[access, 'Directo'] = qd
        df_flows_ade.at[access, 'Izquierda'] = qvi
        df_flows_ade.at[access, 'Derecha'] = qvd

    """
    Notas para el programador:
    Hay casos donde no existe giro a la izquierda o derecha.
    Considera eso en el programa.
    """

    ######################
    # LOST TIME BY CYCLE #
    ######################

    no_group = []
    for index, value in df.iterrows():
        no_group.append(value["Grupo"])

    no_group = list(set(no_group))

    phases_dict = {}
    for no in no_group:
        phases_by_access = []
        for index, value in df.iterrows():
            if value["Grupo"] == no:
                phases_by_access.append(index)
        phases_dict[no] = {
            "Access": phases_by_access,
        }

    ##############################
    # DATOS DE MÍNIMO VERDE Y RR #
    ##############################

    RR = rr_time[rr_time_id]
    MIN_GREEN = min_green[min_green_id][1]

    TOTAL_LOST_TIME = 0
    for _ in range(len(no_group)):
        TOTAL_LOST_TIME += RR + 3 #AMBAR

    #################################################
    # MAX RELATIONS CURRENT FLOW AND SATURATED FLOW #
    #################################################

    max_relations = []
    for phase, value in phases_dict.items():
        maximum_flows_ade = 0
        for (index, row), (index2, row2) in zip(df_flows_ade.iterrows(), df.iterrows()):
            if index in value["Access"]:
                sum_flows = (row["Directo"] + row["Izquierda"] + row["Derecha"])/row2["Carriles"]/1800 #<---- DATO
                if sum_flows > maximum_flows_ade:
                    maximum_flows_ade = sum_flows
        max_relations.append(round(maximum_flows_ade,3))

    ###########################
    # COMPUTING OPTIMAL CYCLE #
    ###########################

    ws = wb_WEBSTER['WEBSTER']

    if sum(max_relations) <= 0.80: 
        Cw = (1.5*TOTAL_LOST_TIME + 5) // (1 - sum(max_relations))
        Cmin = (TOTAL_LOST_TIME)/(1-sum(max_relations))
        Cp = TOTAL_LOST_TIME/(1-1.1*sum(max_relations))
        Ccruce = MIN_GREEN + TOTAL_LOST_TIME
        Cmax = 150 + MIN_GREEN
        ws.cell(17,2).value = "Los flujos no se superan la capacidad, seleccionar el tiempo de ciclo de las opciones propuestas."
    else:
        Cw = 0
        Cmin = 0
        Cp = 0
        Ccruce = MIN_GREEN + TOTAL_LOST_TIME
        Cmax = 150+MIN_GREEN
        ws.cell(17,2).value = "Los flujos superan la capacidad, se recomienda al simulador establecer el Tiempo de Ciclo."

    #TODO: La fila depende de que escenario es :D
    #Tiempo de Ciclo
    ws.cell(2,3).value = Cmax
    ws.cell(2,4).value = Cw
    ws.cell(2,5).value = Cmin
    ws.cell(2,6).value = Cp
    ws.cell(2,7).value = Ccruce

    #Relaciones de flujo y flujo saturado
    for idx, no in enumerate(no_group): #1, 2, 3, ...
        ws.cell(2, idx+10).value = max_relations[idx]

    #Ai y RRi
    for no in no_group:
        ws.cell(2,22+no*3).value = 3
        ws.cell(2,23+no*3).value = RR

    #TODO: Si esta saturado, el microsimulador debe sugerir un tiempo de ciclo máximo.

if __name__ == '__main__':
    path_datos = r"DATOS.xlsx"

    df = pd.read_excel(path_datos, header=0, usecols="A:G", nrows=11).dropna()
    df.index = df.iloc[:,0].astype(int)
    df =df.iloc[:,1:]
    mapping = {'SI': True, 'NO': False}
    df['Protegido'] = df['Protegido'].replace(mapping).infer_objects(copy=False)

    df2 = pd.read_excel(path_datos, header=0, usecols="I:L", nrows=11).dropna()
    rr_time_id = df2.loc[df2['Todo Rojo'].idxmax()]['Caso']
    min_green_id = rr_time_id

    #Paths:
    vehicle_path = Path(r"C:\Users\dacan\OneDrive\Desktop\PRUEBAS\Maxima Entropia\1 PROYECTO SURCO\7. Informacion de Campo\Sub Area 016\Vehicular\Tipico\SS-77_Av. Rosa Lozano-Jr. Geranios_Veh_T.xlsm")
    pedestrian_path = Path(r"C:\Users\dacan\OneDrive\Desktop\PRUEBAS\Maxima Entropia\1 PROYECTO SURCO\7. Informacion de Campo\Sub Area 016\Peatonal\Tipico\SS-77_ Av. Rosa Lozano - Jr. Geranios_Peatonal_T.xlsm")

    subarea_folder = Path(r"C:\Users\dacan\OneDrive\Desktop\PRUEBAS\Maxima Entropia\1 PROYECTO SURCO\6. Sub Area Vissim")
    balance_folder = subarea_folder / "Balanceado"
    intervals = []
    for tipicidad in ["Tipico", "Atipico"]:
        for turno in ["Manana", "Tarde", "Noche"]:
            balance_path = balance_folder / tipicidad / turno / f"Balance_{turno}.xlsx"
            wb = load_workbook(balance_path, read_only=True, data_only=True)
            ws = wb['Sheet1']
            peakhour = ws.cell(2,3).value
            inicio, _ = peakhour.split(" - ")
            hour = (int(inicio[:2]) + int(inicio[3:5])/60)*4
            intervals.append(slice(hour, hour+4))

    """
    0: HPM - TIPICO
    1: HPT - TIPICO
    2: HPN - TIPICO
    3: HPM - ATIPICO
    4: HPT - ATIPICO
    5: HPN - ATIPICO
    """

    for i in range(13):
        #Factores
        if i+1 == 1 or i+1 == 9:
            factor = 0.30
        elif i+1 == 2 or i+1 == 8 or i+1 == 13:
            factor = 0.50
        else:
            factor = 1.00

        #Horas puntas para los intervalos
        if 1 <= i+1 <= 3:
            interval = intervals[0]
        elif i+1 == 4:
            interval = slice(8*4, 9*4)
        elif i+1 == 5:
            interval = intervals[1]
        elif i+1 == 6:
            interval = slice(14*4, 15*4)
        elif 7 <= i+1 <= 8:
            interval = intervals[2]
        elif 9 <= i+1 <= 10:
            interval = intervals[3]
        elif i+1 == 11:
            interval = intervals[4]
        elif 12 <= i+1:
            interval = intervals[5]

    path_template = r".\tools\WEBSTER.xlsx"
    wb_WEBSTER = load_workbook(path_template, read_only=False, data_only=False)

        compute_webster(vehicle_path, pedestrian_path, min_green_id, rr_time_id, interval, df, factor)

    wb_WEBSTER.save("TEST2.xlsx")
    wb_WEBSTER.close()

    #compute_webster(vehicle_path, pedestrian_path, min_green_id, rr_time_id, slice(28, 32), df)
    print("Done")
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import load_workbook
import os
import shutil

dict_tipicity_turn = {
    1: ["Tipico", "HPMAD"],
    2: ["Tipico", "HVMAD"],
    3: ["Tipico", "HPM"],
    4: ["Tipico", "HVM"],
    5: ["Tipico", "HPT"],
    6: ["Tipico", "HVT"],
    7: ["Tipico", "HPN"],
    8: ["Tipico", "HVN"],
    9: ["Atipico", "HVMAD"],
    10: ["Atipico", "HPM"],
    11: ["Atipico", "HPT"],
    12: ["Atipico", "HPN"],
    13: ["Atipico", "HVN"],
}

def _change_sig(
        list_green,
        sig_path,
        cycle_time,
        ) -> None:
    #################################
    # Modifying sigs for each phase #
    #################################

    tree = ET.parse(sig_path)
    sc_tag = tree.getroot()
    interstages = sc_tag.find("./stageProgs/stageProg/interstages")
    for interstage, green in zip(interstages.findall("./interstage"), list_green):
        interstage.attrib["begin"] = str(green*1000)

    stageProg = sc_tag.find("./stageProgs/stageProg")
    stageProg.attrib["cycletime"] = str(int(cycle_time*1000))
    
    ET.indent(tree, "    ")
    tree.write(sig_path, encoding = "utf-8", xml_declaration = True)

def _get_greens(
        TIMES: list, #Route of excel.
        ) -> dict:
    """ Modify sigs for each scenario. """
    ##############################
    # Computing phases and times #
    ##############################

    programs_dict = {}
    for index, row in enumerate(TIMES):
        program = []
        for i in range(len(row)):
            if i%3 == 0:
                program.append(row[i:i+3])
        programs_dict[index+1] = program
        
    program_0 = programs_dict[1]
    for i, phases in enumerate(program_0):
        try:
            if sum(phases) == 0: #TODO: Hacer que todo tenga 0 =D
                no_phases = i
                break
        except TypeError:
            no_phases = i
            break

    ###################################
    # Computing greens times for sigs #
    ###################################

    begin_dict = {}
    cycle_dict = {}
    for key, value in programs_dict.items():
        row = value[:no_phases]
        greens = []
        aux = 0
        for i in range(len(row)):
            if i==0:
                greens.append(row[i][0])
            else:
                aux = aux + sum(row[i-1])
                greens.append(aux+row[i][0])

        begin_dict[key] = greens
        cycleTime = 0
        for i in range(len(row)):
            cycleTime += sum(row[i])

        cycle_dict[key] = cycleTime

    return begin_dict, cycle_dict


def replicate_sigs(subareaPath): #OK
    proposedPath = os.path.join(subareaPath, "Propuesto")

    patternPath = os.path.join(proposedPath, "Tipico", "HPM")
    listSigPatterns = os.listdir(patternPath)
    listSigPatterns = [file for file in listSigPatterns if file.endswith('.sig')]
    listSigCodes = [file[:-4] for file in listSigPatterns]
    listSigPatternsPath = [os.path.join(patternPath, file) for file in listSigPatterns]

    #Getting greens and cycle times
    programResultPath = os.path.join(subareaPath, "Program_Results.xlsx")
    wb = load_workbook(programResultPath, read_only=True, data_only=True)

    dataSubarea = {}

    for code in listSigCodes:
        ws = wb[code]
        TIMES = [[elem.value for elem in row] for row in ws[slice("V2","AJ14")]]
        beginDict, cycleDict = _get_greens(TIMES)
        dataSubarea[code] = [beginDict, cycleDict]

    wb.close()

    for tipicidad in ["Tipico", "Atipico"]:
        tipicidadPath  = os.path.join(proposedPath, tipicidad)
        turnos = os.listdir(tipicidadPath)
        turnos = [file for file in turnos if not file.endswith('.ini')]

        for turno in turnos:
            if turno == "HPM" and tipicidad == "Tipico": continue
            if not turno in ["HPM", "HPT", "HPN"]: continue
            turnoPath = os.path.join(tipicidadPath, turno)
            for sigOriginPath, sigName in zip(listSigPatternsPath, listSigPatterns):
                sigFinalPath = os.path.join(turnoPath, sigName)
                shutil.copy2(sigOriginPath, sigFinalPath)

    propuesto_folder = Path(subareaPath) / "Propuesto"
    for tipicidad in ["Tipico", "Atipico"]:
        print(f"\n{f' Tipicidad: {tipicidad} ':#^{50}}")
        list_files = os.listdir(propuesto_folder / tipicidad)
        list_files = [file for file in list_files if not file.endswith(".ini")]
        for scenario in list_files:
            print(f"Escenario: {scenario}")
            listFiles = os.listdir(propuesto_folder / tipicidad / scenario)
            sig_files = [file for file in listFiles if file.endswith(".sig")]
            for sig in sig_files:
                for index, listDuo in dict_tipicity_turn.items():
                    if scenario == listDuo[1] and tipicidad == listDuo[0]:
                        for codeNode, listData in dataSubarea.items():
                            if sig[:-4] == codeNode:
                                begin_dict, cycle_dict = listData
                                sig_path = propuesto_folder / tipicidad / scenario / sig
                                _change_sig(begin_dict[index], sig_path, cycle_dict[index])
                                #supplyFile2
                                break
            inpx_files = [file for file in listFiles if file.endswith(".inpx")]
            if len(inpx_files) > 1: raise Exception("More than one inpx file")

            inpx_file = inpx_files[0]
            inpxPath = propuesto_folder / tipicidad / scenario / inpx_file
            tree = ET.parse(inpxPath)
            networkTag = tree.getroot()

            for signalController in networkTag.findall("./signalControllers/signalController"):
                name = signalController.get('name')
                signalController.set("supplyFile2", f"./{name}.sig")

            ET.indent(tree)
            tree.write(inpxPath, encoding = "utf-8", xml_declaration = True)

    print(f"{' Copia y modificación finalizado ':#^{50}}")
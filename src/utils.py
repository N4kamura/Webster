import xml.etree.ElementTree as ET
import os
from openpyxl import load_workbook
import re
import numpy as np
import pandas as pd
import win32com.client as com
from dataclasses import dataclass
from pathlib import Path
import xlsxwriter
import xlsxwriter.format
import xlsxwriter.worksheet
import xlwings as xw
from tqdm import tqdm
import pywintypes

HEADERS = [
                "Escenario", "Cf", "Cmax", "Cw", "Cmin", "Cp", "Ccruce", "L", "Lpeat", "gT",
                "Y1", "Y2", "Y3", "Y4", "Y5", "g1", "g2", "g3", "g4", "g5", "CHECK",
                "V1", "A1", "RR1",
                "V2", "A2", "RR2",
                "V3", "A3", "RR3",
                "V4", "A4", "RR4",
                "V5", "A5", "RR5",
                "Vmin1", "Vmin2", "Vmin3", "Vmin4", "Vmin5",
            ]

SCENARIO_NAMES = [
    "1: HPMAD", "2: HVMAD", "3:HPM", "4: HVM", "5: HPT", "6: HVT", "7: HPN", "8: HVN",
    "9: HVMAD", "10: HPM", "11:HPT", "12: HPN", "13: HVN",
]

def process_elem(item: str | int) -> list:
    if isinstance(item, str):
        return list(map(int, item.split(',')))
    else:
        return [item]

def get_codes(subareaPath, error_message):
    listContent = os.listdir(subareaPath)

    #Filters to obtain .inpx file (skeleton)
    listContent = [file for file in listContent if file.endswith(".inpx") and "(SA)" in file]

    #Check if only one .inpx file is found
    if len(listContent) > 1:
        return error_message.showMessage("More than one .inpx file found")

    vissimFile = listContent[0]
    vissimPath = os.path.join(subareaPath, vissimFile)

    #Obtaining codes from inpx file (skeleton)
    tree = ET.parse(vissimPath)
    network_tag = tree.getroot()
    listCodes = []
    for node_tag in network_tag.findall("./nodes/node"):
        for uda_tag in node_tag.findall("./uda"):
            code = uda_tag.get("value")
            listCodes.append(code)

    tree = None

    return listCodes

def _get_interval_from_excel(excelPath) -> slice:
    try:
        wb = load_workbook(excelPath, read_only=True, data_only=True)
    except Exception as e:
        print("Error en:\n", excelPath)
        return None
    ws = wb['Histograma']
    intervals = []
    for j in range(3):
        peakHour = ws.cell(18+j*7,3).value
        hour = int((int(peakHour[:2]) + int(peakHour[3:5])/60)*4)
        intervals.append(slice(hour, hour + 4))
    
    wb.close()
    
    return intervals

def get_dict_by_agent(subareaFolder: str, excel_by_agent: dict, selectedCode: str) -> list[dict, list[slice]]:
    pattern = r"([A-Z]+-[0-9]+)"
    intervals = []
    for agentName in ["Vehicular", "Peatonal"]:
        for tipicidad in ["Tipico", "Atipico"]:
            agentFolder = os.path.join(
                subareaFolder,
                agentName,
                tipicidad,
            )

            agentContentFolder = os.listdir(agentFolder)
            for agentFile in agentContentFolder:
                if re.search(pattern, agentFile):
                    codeFound = re.search(pattern, agentFile).group(1)
                    if selectedCode == codeFound:
                        agentFilePath = os.path.join(
                            agentFolder,
                            agentFile,
                        )
                        if agentName == "Vehicular":
                            intervalObtained = _get_interval_from_excel(agentFilePath)
                            intervals.extend(intervalObtained)
                        excel_by_agent[agentName][tipicidad] = agentFilePath
                        break

    return excel_by_agent, intervals

def flows(vehicle_types: list,
          wb: load_workbook,
          FACTOR: float,
          interval: slice,
          ) -> np.array:
    origin_slices = [
        slice("E12", "E21"),
        slice("K12", "K21"),
        slice("E24", "E33"),
        slice("K24", "K33"),
    ]

    destiny_slices = [
        slice("F12", "F21"),
        slice("L12", "L21"),
        slice("F24", "F33"),
        slice("L24", "L33"),
    ]

    giro_slices = [
        slice("G12", "G21"),
        slice("M12", "M21"),
        slice("G24", "G33"),
        slice("M24", "M33"),
    ]

    num_giros = [0,0,0,0]
    num_veh_classes = len(vehicle_types)

    ws = wb['Inicio']
    for j, turn in enumerate(giro_slices):
        aux = []
        for row in ws[turn]:
            aux.append(row[0].value)
        try:
            quant = aux.index(None)
        except ValueError:
            quant = len(aux)
        num_giros[j] = quant

    hojas = ["N", "S", "E", "O"]

    list_destination    = []
    list_origin         = []
    list_vr_name        = []

    list_flow = []
    for i_giro in range(len(hojas)):
        ws = wb["Inicio"]
        slice_origin = origin_slices[i_giro]
        slice_destiny = destiny_slices[i_giro]
        slice_giros = giro_slices[i_giro]
        num_giro_i = num_giros[i_giro]

        list_origin.extend(
            [row[0].value for row in ws[slice_origin]][:num_giro_i]
        )

        list_destination.extend(
            [row[0].value for row in ws[slice_destiny]][:num_giro_i]
        )

        list_vr_name.extend(
            [row[0].value for row in ws[slice_giros]][:num_giro_i]
        )

        ws = wb[hojas[i_giro]]

        list_A = [[cell.value for cell in row] for row in ws["K16":"HB111"]]
        A = np.array(list_A, dtype="float")
        A[np.isnan(A)] = 0
        A = A*FACTOR

        list_flow.append(
            np.array(
                [
                    A[interval, (10*veh_type):(10*veh_type+num_giro_i)]
                    for veh_type in range(num_veh_classes)
                ]
            )
        )

        array_flow = np.concatenate(list_flow, axis=-1)

    return array_flow, list_origin, list_destination, list_vr_name

def pedestrian_flows(wb: load_workbook, interval: slice):
    ws = wb['Inicio']

    movimiento_slices = [
        slice("G13", "G22"),
        slice("K13", "K22"),
        slice("G25", "G34"),
        slice("K25", "K34"),
    ]

    types_ped_slices = [
        slice("W4", "W8"),
        slice("Y4", "Y8"),
        slice("AA4", "AA8"),
    ]

    num_ped_classes = 0
    for type_ped_slice in types_ped_slices:
        num_ped_classes += [row[0].value for row in ws[type_ped_slice]].index(None)

    num_giros = [0,0,0,0]
    for j, turn in enumerate(movimiento_slices):
        aux = []
        for row in ws[turn]:
            aux.append(row[0].value)
        try: quant = aux.index(None)
        except ValueError: quant = len(aux)
        num_giros[j] = quant

    list_movimientos = []
    list_ped = []

    for i_giro in range(len(num_giros)):
        ws = wb['Inicio']
        slice_movimiento = movimiento_slices[i_giro]
        num_giro_i = num_giros[i_giro]

        list_movimientos.extend([row[0].value for row in ws[slice_movimiento]][:num_giro_i])

        ws = wb['Data Peatonal']

        list_A = [[cell.value for cell in row] for row in ws["L20":"UY83"]]
        A = np.array(list_A, dtype="float")
        A[np.isnan(A)] = 0
        A = np.concatenate((np.zeros((24, A.shape[1])), A), axis = 0)

        list_ped.append(
            np.array(
                [
                    A[interval, (10*ped_type + 140*i_giro):(10*ped_type + num_giro_i + 140*i_giro)]
                    for ped_type in range(num_ped_classes)
                ]
            )
        )

        array_ped_flow = np.concatenate(list_ped, axis=-1)

        str_movements = [str(no) for no in list_movimientos]

    list_sums = []
    for ped_type in range(num_ped_classes):
        check_status = [False for _ in range(len(str_movements))]
        ped_flow = array_ped_flow[ped_type]
        list_sum_peds = []
        for i, giro in enumerate(str_movements):
            flow_cross = 0
            if check_status[i]: continue
            check_status[i] = True
            flow_cross += sum(ped_flow[:,i])
            for j, giro_inv in enumerate(str_movements):
                if giro[::-1] == giro_inv:
                    check_status[j] = True
                    flow_cross += sum(ped_flow[:,j])
                    break
            list_sum_peds.append(flow_cross)
        list_sums.append(list_sum_peds)

    sum_by_peds = [sum(x) for x in zip(*list_sums)]

    max_ped_flow = max(sum_by_peds)

    return max_ped_flow

def compute_flows(origin: int, dfTurns: pd.DataFrame, direction: int, dfFlows: pd.DataFrame, array_flow):
    flow = 0
    listTurn = dfTurns[(dfTurns["Origen"] == origin) & (dfTurns["Giro"] == direction)].index.tolist()
    for leftTurnIndex in listTurn:
        for veh_type in range(len(array_flow)):
            try:
                flow += sum(array_flow[veh_type][:,leftTurnIndex])
            except IndexError as e:
                print("Posiblemente no existe un giro en el típico o el atípico.")
    if flow == None:
        dfFlows.at[origin, direction] = 0
    else:
        dfFlows.at[origin, direction] = flow

def data2excel(subareaFolder: str, destiny_route: str) -> None:
    """ Create an excel with data from counting files. """
    #Find field files:
    pathDirectory = Path(subareaFolder)
    subareaName = pathDirectory.name
    projectPath = pathDirectory.parents[1] #Va alrevés
    fieldPath = projectPath / "7. Informacion de Campo" / subareaName / "Vehicular"
    typicalPath = fieldPath / "Tipico"
    excelVehicles = os.listdir(typicalPath)
    excelVehicles = [file for file in excelVehicles if file.endswith(".xlsm") and not file.startswith("~$")]
    pattern = r"([A-Z]+-[0-9]+)"

    @dataclass
    class excelData:
        origins: list
        destinations: list
        names: list

    dictInfo = {}
    for excel in tqdm(excelVehicles, desc="Procesando excels"):
        excelPath = typicalPath / excel
        code = re.search(pattern, excel).group(1)
        wb = load_workbook(excelPath, read_only=True, data_only=True)
        ws = wb["Inicio"]
        slicesOrigin = [
            slice("E12", "E22"),
            slice("K12", "K22"),
            slice("E24", "E34"),
            slice("K24", "K34"),
        ]

        slicesDestination = [
            slice("F12", "F22"),
            slice("L12", "L22"),
            slice("F24", "F34"),
            slice("L24", "L34"),
        ]

        slicesNames = [
            slice("G12", "G22"),
            slice("M12", "M22"),
            slice("G24", "G34"),
            slice("M24", "M34"),
        ]
        
        listOrigins = []
        listDestinations = []
        listNames = []

        for sliceOrigin, sliceDestination, sliceName in zip(slicesOrigin, slicesDestination, slicesNames):
            content = [row[0].value for row in ws[sliceOrigin] if row[0].value is not None]
            listOrigins.extend(content)
            content = [row[0].value for row in ws[sliceDestination] if row[0].value is not None]
            listDestinations.extend(content)
            content = [row[0].value for row in ws[sliceName] if row[0].value is not None]
            listNames.extend(content)

        wb.close()

        data = excelData(
            origins = listOrigins,
            destinations = listDestinations,
            names = listNames
        )

        dictInfo[code] = data

    #Writing data in excel

    try:
        excel = com.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
    except Exception as inst:
        print(str(inst))

    try:
        wb = excel.Workbooks.Open(destiny_route)
    except pywintypes.com_error as inst:
        raise inst


    for nameSheet, dataList in tqdm(dictInfo.items(), desc="Escribiendo hojas"):
        try:
            ws = wb.Sheets[nameSheet]
        except pywintypes.com_error as inst:
            print("\nNo existe el sheet: ", nameSheet)
            print(str(inst))
            continue
        except Exception as inst:
            print("\nError: ", nameSheet)
            print(str(inst))
            
        try:
            for i in range(len(dataList.origins)):
                ws.Cells(29+i, 1).Value = dataList.origins[i]
                ws.Cells(29+i, 2).Value = dataList.destinations[i]
                ws.Cells(29+i, 4).Value = dataList.names[i]

            originList = list(set(dataList.origins))
            
            for j, origin in enumerate(originList):
                ws.Cells(29+j, 10).Value = origin
        except Exception as inst:
            print("\nError: ", nameSheet)
            print(str(inst))
            continue

    wb.Save()
    wb.Close()

def _config_excel(worksheet: xlsxwriter.worksheet, cell_format: xlsxwriter.format, cell_format2: xlsxwriter.format, cell_format3: xlsxwriter.format) -> None:
    for row in range(14):
        for col in range(41):
            worksheet.write(row, col, '', cell_format)
    
    for i, header in enumerate(HEADERS):
        worksheet.write(0, i, header, cell_format3)

    for i, scenarioName in enumerate(SCENARIO_NAMES):
        "A1:AO14"
        worksheet.write(1+i, 0, scenarioName, cell_format2)

    for i in range(2,15):
        worksheet.write_formula(i-1, 7, f"=SUM(W{i}:X{i})+SUM(Z{i}:AA{i})+SUM(AC{i}:AD{i})+SUM(AF{i}:AG{i})+SUM(AI{i}:AJ{i})+I{i}", cell_format)
        worksheet.write_formula(i-1, 9, f"=B{i}-H{i}", cell_format)
        worksheet.write_formula(i-1, 15, f"=IFERROR(K{i}/SUM($K{i}:$O{i})*$J{i},0)", cell_format)
        worksheet.write_formula(i-1, 16, f"=IFERROR(L{i}/SUM($K{i}:$O{i})*$J{i},0)", cell_format)
        worksheet.write_formula(i-1, 17, f"=IFERROR(M{i}/SUM($K{i}:$O{i})*$J{i},0)",cell_format)
        worksheet.write_formula(i-1, 18, f"=IFERROR(N{i}/SUM($K{i}:$O{i})*$J{i},0)", cell_format)
        worksheet.write_formula(i-1, 19, f"=IFERROR(O{i}/SUM($K{i}:$O{i})*$J{i},0)", cell_format)

        worksheet.write_formula(i-1, 20, f'=IF(SUM(V{i}:AJ{i})=B{i}, IF(SUM(P{i}:T{i})=J{i}, "OK", "IMBALANCE"), IF(SUM(V{i}:AJ{i})>B{i}, CONCAT(" - ", ROUND(SUM(V{i}:AJ{i})-B{i}, 0)), CONCAT(" + ", ROUND(B{i}-SUM(V{i}:AJ{i}), 0))))', cell_format)

        worksheet.write_formula(i-1, 21, f"=IF(P{i}<AK{i}, AK{i}, P{i})", cell_format)
        worksheet.write_formula(i-1, 24, f"=IF(Q{i}<AL{i}, AL{i}, Q{i})", cell_format)
        worksheet.write_formula(i-1, 27, f"=IF(R{i}<AM{i}, AM{i}, R{i})", cell_format)
        worksheet.write_formula(i-1, 30, f"=IF(S{i}<AN{i}, AN{i}, S{i})", cell_format)
        worksheet.write_formula(i-1, 33, f"=IF(T{i}<AO{i}, AO{i}, T{i})", cell_format)

    worksheet.set_column('I:T', None, None, {'hidden': True})

def duplicate_name_sheets(listCodes: list, finalPath: str) -> None:
    workbook = xlsxwriter.Workbook(finalPath)
    cell_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
    })
    
    cell_format2 = workbook.add_format({
        'border': 1,
        'align': 'left',
        'valign': 'vcenter',
    })

    cell_format3 = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'bold': True
    })

    for code in listCodes:
        worksheet = workbook.add_worksheet(code)
        _config_excel(worksheet, cell_format, cell_format2, cell_format3)

    workbook.close()

def duplicate2(listCodes: list, finalPath: str) -> None:
    app = xw.App(visible=False)
    wb = app.books.open(finalPath)

    source_sheet = wb.sheets["DATA"]
    for i, code in enumerate(tqdm(listCodes, desc="Copiando hojas")):
        source_sheet.api.Copy(After = source_sheet.api)

        new_sheet = wb.sheets[-1-i]
        new_sheet.name = code

    wb.sheets["DATA"].delete()

    wb.save()
    wb.close()
    app.quit()